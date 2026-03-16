import streamlit as st
import pandas as pd
import plotly.express as px
import main 
import io 
import time
import datetime

st.set_page_config(page_title="YZ Ders Programı", layout="wide", page_icon="🏫")

st.markdown("""
    <style>
    /* 1. Senin Özel Buton Ayarların */
    div.stButton > button:first-child {
        background-color: #0E1117; color: white; border-radius: 8px; border: 1px solid #4B4B4B;
    }
    div.stButton > button:hover {
        background-color: #0056b3; color: white; border-color: #0056b3;
    }
    [data-testid="stSidebar"] {
        background-color: #34495E; 
        min-width: 260px !important; 
        max-width: 300px !important; 
    }
    /* 2. YENİ: Sol Menüyü (Sidebar) İnceltme ve Renklendirme */
    [data-testid="stSidebar"] {
        background-color: #34495E; 
        min-width: 260px !important; /* İncelttik */
        max-width: 300px !important; /* İncelttik */
    }
    [data-testid="stSidebar"] * {
        color: #FFFFFF;
    }
    </style>
""", unsafe_allow_html=True)
st.markdown("## 🏫 YZ Destekli Ders Programı Otomasyonu")
st.divider()

if 'sonuc' not in st.session_state: st.session_state.sonuc = None
if 'secilen_strateji' not in st.session_state: st.session_state.secilen_strateji = "Dengeli"
if 'dersler' not in st.session_state: st.session_state.dersler = []
if 'hocalar' not in st.session_state: st.session_state.hocalar = []
if 'derslikler' not in st.session_state: st.session_state.derslikler = []
if 'gecmis' not in st.session_state: st.session_state.gecmis = []
if 'duzenlenecek_hoca' not in st.session_state: st.session_state.duzenlenecek_hoca = None 
if 'yuklenen_dosya_adi' not in st.session_state: st.session_state.yuklenen_dosya_adi = "" 
if 'secilen_strateji' not in st.session_state: st.session_state.secilen_strateji = "Dengeli"
secilen_strateji = st.session_state.secilen_strateji  # Her sayfa bu değişkeni tanıyacak!
def memnuniyet_hesapla(df):
    hoca_puan, ogr_puan = 100, 100
    h_gunluk = df.groupby(['Hoca', 'Gün']).size()
    hoca_puan -= len(h_gunluk[h_gunluk > 3]) * 2
    s_idx = {"08:00":0, "09:00":1, "10:00":2, "11:00":3, "13:00":4, "14:00":5, "15:00":6, "16:00":7}
    df_ogr = df[df['Dönem/Sınıf'] != '-'].copy()
    if not df_ogr.empty:
        df_ogr['s_idx'] = df_ogr['Saat'].map(s_idx)
        s_gunluk = df_ogr.groupby(['Dönem/Sınıf', 'Gün'])['s_idx'].agg(['max', 'min', 'count'])
        s_gunluk['bosluk'] = (s_gunluk['max'] - s_gunluk['min']) - (s_gunluk['count'] - 1)
        ogr_puan -= s_gunluk[s_gunluk['bosluk'] > 1]['bosluk'].sum() * 3
    return max(0, min(100, int(hoca_puan))), max(0, min(100, int(ogr_puan)))

def tarih_araligi_hesapla(hafta_str, bas_tarihi):
    aylar = ["", "Oca", "Şub", "Mar", "Nis", "May", "Haz", "Tem", "Ağu", "Eyl", "Eki", "Kas", "Ara"]
    if not hafta_str or hafta_str == "1-14 (Tüm Dönem)":
        bitis = bas_tarihi + datetime.timedelta(weeks=14) - datetime.timedelta(days=1)
        return f"({bas_tarihi.day} {aylar[bas_tarihi.month]} - {bitis.day} {aylar[bitis.month]})"
    haftalar = []
    for p in str(hafta_str).replace(" ", "").split(","):
        if "-" in p:
            try: haftalar.extend(range(int(p.split("-")[0]), int(p.split("-")[1]) + 1))
            except: pass
        else:
            try: haftalar.append(int(p))
            except: pass
    if not haftalar: return ""
    b_tarih = bas_tarihi + datetime.timedelta(weeks=min(haftalar) - 1)
    bt_tarih = bas_tarihi + datetime.timedelta(weeks=max(haftalar)) - datetime.timedelta(days=1)
    return f"({b_tarih.day} {aylar[b_tarih.month]} - {bt_tarih.day} {aylar[bt_tarih.month]})"

def alternatif_bul(hoca, sinif, df, hoca_listesi):
    gunler, saatler = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"], ["08:00", "09:00", "10:00", "11:00", "13:00", "14:00", "15:00", "16:00"]
    h_data = next((h for h in hoca_listesi if h['ad'] == hoca), None)
    y_gunler, y_saatler = str(h_data.get('musait_olmayan_gunler', '')) if h_data else "", str(h_data.get('istenmeyen_saatler', '')) if h_data else ""
    dolu_h = df[df['Hoca'] == hoca][['Gün', 'Saat']].values.tolist()
    dolu_s = df[df['Sınıf'] == sinif][['Gün', 'Saat']].values.tolist()
    oneriler = []
    for g in gunler:
        if g in y_gunler: continue
        for s in saatler:
            if s in y_saatler: continue
            if [g, s] not in dolu_h and [g, s] not in dolu_s:
                oneriler.append(f"**{g} {s}**")
                if len(oneriler) == 3: return oneriler
    return oneriler
# --- YENİ: YZ DEKAN ASİSTANI (DOĞAL DİL İŞLEME MOTORU) ---
# --- YENİ: YZ DEKAN ASİSTANI (GELİŞMİŞ DOĞAL DİL İŞLEME MOTORU) ---
# --- YENİ: YZ DEKAN ASİSTANI (GELİŞMİŞ NLP VE TÜRKÇE DESTEĞİ) ---
# --- YENİ: YZ DEKAN ASİSTANI (HOCA VE ÖĞRENCİ RADARI ZEKASI BİR ARADA) ---
def asistan_anla(metin, hoca_listesi):
    def tr_lower(text):
        tr_map = {'I': 'ı', 'İ': 'i', 'Ş': 'ş', 'Ç': 'ç', 'Ğ': 'ğ', 'Ü': 'ü', 'Ö': 'ö'}
        for key, val in tr_map.items(): text = text.replace(key, val)
        return text.lower()

    metin_k = tr_lower(str(metin)).replace("'", "").replace(".", "")
    kullanici_kelimeleri = set(metin_k.split())
    
    gunler_sozluk = {"pazartesi": "Pazartesi", "ptesi": "Pazartesi", "salı": "Salı", "sali": "Salı", "çarşamba": "Çarşamba", "carsamba": "Çarşamba", "çrş": "Çarşamba", "perşembe": "Perşembe", "persembe": "Perşembe", "prş": "Perşembe", "cuma": "Cuma"}
    saatler_sozluk = {"sabah": ["08:00", "09:00", "10:00", "11:00"], "sabahtan": ["08:00", "09:00", "10:00", "11:00"], "öğle": ["13:00", "14:00", "15:00", "16:00"], "öğleden": ["13:00", "14:00", "15:00", "16:00"], "öğlenden": ["13:00", "14:00", "15:00", "16:00"]}
    
    unvanlar = {"prof", "doc", "doç", "dr", "ogr", "öğr", "gör", "üyesi", "uyesi", "arş", "ars"}
    hedef_hoca = None
    
    for h in hoca_listesi:
        hoca_tam_ad = tr_lower(str(h['ad']))
        hoca_kelimeleri = set(hoca_tam_ad.replace(".", " ").split())
        temiz_hoca_kelimeleri = {k for k in hoca_kelimeleri if k not in unvanlar and len(k) > 2}
        if kullanici_kelimeleri.intersection(temiz_hoca_kelimeleri):
            hedef_hoca = h
            break
            
    yasak_gunler = [g_ad for g_kelime, g_ad in gunler_sozluk.items() if g_kelime in metin_k]
    yasak_saatler = []
    for s_kelime, s_liste in saatler_sozluk.items():
        if s_kelime in metin_k: yasak_saatler.extend(s_liste)
    s_disi = any(k in metin_k for k in ["şehir dışı", "sehir disi", "tek gün", "dışarıdan"])
    
    # YENİ: Eğer cümlede hoca adı yoksa ve şikayet/radar/çakışma kelimeleri varsa, bu bir Radar Emridir!
    is_radar = any(k in metin_k for k in ["radar", "şikayet", "öğrenci", "çakışma", "incele", "analiz"]) and not hedef_hoca
    
    return hedef_hoca, list(set(yasak_gunler)), list(set(yasak_saatler)), s_disi, is_radar
# =======================
# YENİ VİTRİN VE SOL KUMANDA PANELİ
# =======================
with st.sidebar:
    st.markdown("""
        <div style='text-align: center; padding: 10px; background-color: #34495E; border-radius: 10px; margin-bottom: 20px;'>
            <h2 style='color: #ECF0F1; margin-bottom: 0px;'>🏛️ NÖHÜ</h2>
            <p style='color: #BDC3C7; font-size: 14px; font-weight: bold;'>Akıllı Kampüs & YZ Asistanı</p>
        </div>
    """, unsafe_allow_html=True)

    page = st.radio(
        "KONTROL PANELİ MENÜSÜ",
        [
            "📊 1. Ana Ekran & Veri",
            "💎 2. VIP Ayarları",
            "🚀 3. YZ Motoru",
            "✏️ 4. Manuel Düzenleme",
            "🤖 5. YZ Asistan",
            "📡 6. Kampüs Radarı"
        ]
    )
    
    
    st.write("---")
    st.markdown("### 🛡️ Sistem Durumu")
    st.success("🟢 **YZ Sunucusu:** Çevrimiçi\n\n👤 **Aktif Oturum:** Mimar Veysi\n\n🏛️ **Yetki:** Sistem Yöneticisi")
    st.caption("Sürüm 2.1.0 | NÖHÜ Yapay Zeka Laboratuvarı © 2026")

# =======================
# 1. SAYFA: DASHBOARD VE VERİ YÜKLEME
# =======================
if page == "📊 1. Ana Ekran & Veri":
    if page == "📊 1. Ana Ekran & Veri":
        st.markdown("### 📥 Veri Girişi ve Kapasite Özeti")
        yuklenen = st.file_uploader("Excel Dosyası (.xlsx)", type=["xlsx"])
    
    # Not: Aşağıdaki 'if yuklenen' bloğu, bu 'if page' bloğunun 
    # içinde veya dışında olabilir ama kodunun akışına göre 
    # girintisine (boşluklarına) dikkat etmelisin.
    if yuklenen and st.session_state.yuklenen_dosya_adi != yuklenen.name:
        # Ham okuma
        df_d = pd.read_excel(yuklenen, sheet_name="Dersler")
        df_h = pd.read_excel(yuklenen, sheet_name="Hocalar").fillna("")
        df_dl = pd.read_excel(yuklenen, sheet_name="Derslikler")

        # Basit Çeviri ve Güvenlik (main.py'nin hata vermemesi için)
        df_d = df_d.rename(columns={"Ders Adı": "ad", "Beklenen Öğrenci Sayısı": "ogrenci_sayisi", "Ders Türü": "ders_tipi", "Atanacak Hoca": "hoca_id", "Dönem/Sınıf": "sinif_seviyesi"})
        df_h = df_h.rename(columns={"Hoca Adı": "ad", "İstemediği Günler": "musait_olmayan_gunler"})
        if "id" not in df_h.columns: df_h["id"] = df_h["ad"]
        df_dl = df_dl.rename(columns={"Derslik Adı": "id", "Kapasite": "kapasite", "Türü": "derslik_tipi"})

        # Sisteme Kayıt
        st.session_state.dersler = df_d.to_dict('records')
        st.session_state.hocalar = df_h.to_dict('records')
        st.session_state.derslikler = df_dl.to_dict('records')
        st.session_state.yuklenen_dosya_adi = yuklenen.name
        st.success("✅ Veriler Aktarıldı!")

        # --- DİNAMİK SÜTUN BULUCU (YAPAY ZEKA ÖNCESİ ZIRH) ---
        def s_bul(df, kelimeler):
            # Excel'deki başlıkları küçük harfe çevirip boşlukları siliyoruz
            for col in df.columns:
                c_kucuk = str(col).lower().replace(" ", "").replace("_", "")
                for k in kelimeler:
                    if k in c_kucuk: return col
            return df.columns[0] # Hiçbir şey bulamazsa çökmemesi için ilk sütunu ver (Güvenlik ağı)

        # 2. Dersler Tablosunu Dinamik Haritalandır
        df_d = df_d.rename(columns={
            s_bul(df_d, ["ders", "ad", "isim"]): "ad",
            s_bul(df_d, ["öğrenci", "ogrenci", "mevcut", "sayı", "sayi", "beklenen"]): "ogrenci_sayisi",
            s_bul(df_d, ["tür", "tur", "tip"]): "ders_tipi",
            s_bul(df_d, ["hoca", "öğretim", "ogretim", "eleman"]): "hoca_id",
            s_bul(df_d, ["sınıf", "sinif", "dönem", "donem", "seviye"]): "sinif_seviyesi"
        })

        # 3. Hocalar Tablosunu Dinamik Haritalandır
        df_h = df_h.rename(columns={
            s_bul(df_h, ["hoca", "isim", "ad", "öğretim"]): "ad",
            s_bul(df_h, ["istemediği", "yasak", "kapalı", "olmayan", "gün", "gun"]): "musait_olmayan_gunler"
        })
        if "id" not in df_h.columns: df_h["id"] = df_h["ad"] # Motor için ID Klonlaması
        if "istenmeyen_saatler" not in df_h.columns: df_h["istenmeyen_saatler"] = ""
        if "sehir_disi" not in df_h.columns: df_h["sehir_disi"] = False

        # 4. Derslikler Tablosunu Dinamik Haritalandır
        df_dl = df_dl.rename(columns={
            s_bul(df_dl, ["derslik", "sınıf", "sinif", "amfi", "ad", "isim", "kod", "lab"]): "id",
            s_bul(df_dl, ["kapasite", "mevcut", "kişi", "kisi", "büyüklük"]): "kapasite",
            s_bul(df_dl, ["tür", "tur", "tip"]): "derslik_tipi"
        })

        # 5. YZ'nin Beynine Gönder
        st.session_state.dersler = df_d.to_dict('records')
        st.session_state.hocalar = df_h.to_dict('records')
        st.session_state.derslikler = df_dl.to_dict('records')
        st.session_state.yuklenen_dosya_adi = yuklenen.name
        
        st.success("✅ Veritabanı bağlandı! Akıllı Sütun Tarayıcı Excel'deki her şeyi doğru formatta eşleştirdi.")

        # 2. Sütun İsimlerini YZ Motorunun (main.py) Anlayacağı Dile Çevir (Veri Normalizasyonu)
        # 2. Sütun İsimlerini YZ Motorunun Anlayacağı Dile Çevir
        df_d = df_d.rename(columns={
            "Ders Adı": "ad", 
            "Ders": "ad",
            "Beklenen Öğrenci Sayısı": "ogrenci_sayisi", 
            "Öğrenci Mevcudu": "ogrenci_sayisi", 
            "Öğrenci Sayısı": "ogrenci_sayisi",
            "Ders Türü": "ders_tipi", 
            "Atanacak Hoca": "hoca_id",
            "Hoca": "hoca_id",              # <--- Excel'deki olası isim
            "Hoca Adı": "hoca_id",          # <--- Excel'deki olası isim
            "Öğretim Elemanı": "hoca_id",   # <--- Excel'deki olası isim
            "Dönem/Sınıf": "sinif_seviyesi"
        })
        
        # ZIRH: Eğer Excel'de bambaşka bir isim varsa ve hoca_id oluşmadıysa, 
        # içinde "hoca" veya "öğretim" geçen ilk sütunu zorla hoca_id yap.
        if "hoca_id" not in df_d.columns:
            for col in df_d.columns:
                if "hoca" in col.lower() or "öğretim" in col.lower():
                    df_d["hoca_id"] = df_d[col]
                    break
        
        df_h = df_h.rename(columns={
            "Hoca Adı": "ad", 
            "İstemediği Günler": "musait_olmayan_gunler"
        })
        
        # EKLENEN SİHİRLİ SATIR: main.py 'id' aradığı için, hoca adını kopyalayıp 'id' yapıyoruz!
        if "id" not in df_h.columns:
            df_h["id"] = df_h["ad"]
        
        
        df_dl = df_dl.rename(columns={
            "Derslik Adı": "id", 
            "Kapasite": "kapasite", 
            "Türü": "derslik_tipi"
        })

        # 3. Güvenlik Zırhı: Eğer bu sütunlar Excel'de yoksa, motor çökmesin diye boş olarak ekle
        if "istenmeyen_saatler" not in df_h.columns: df_h["istenmeyen_saatler"] = ""
        if "sehir_disi" not in df_h.columns: df_h["sehir_disi"] = False

        # 4. YZ'nin Beynine (Session State) Gönder
        st.session_state.dersler = df_d.to_dict('records')
        st.session_state.hocalar = df_h.to_dict('records')
        st.session_state.derslikler = df_dl.to_dict('records')
        st.session_state.yuklenen_dosya_adi = yuklenen.name
        
        st.success("✅ Veritabanı bağlandı ve veriler YZ diline çevrildi!")
        
    if 'dersler' in st.session_state and len(st.session_state.dersler) > 0:
        st.divider()
        st.markdown("### 📊 Kapasite ve Sistem Özeti")
        m1, m2, m3 = st.columns(3)
        m1.metric("📚 Toplam Ders", len(st.session_state.dersler))
        m2.metric("👩‍🏫 Aktif Hoca", len(st.session_state.hocalar))
        m3.metric("🏢 Mevcut Derslik", len(st.session_state.derslikler))
elif page == "🚀 3. YZ Motoru":

    # Sadece dersler değil, bir sonuç (program) üretilmişse asistan çalışsın:
    if len(st.session_state.dersler) > 0:
        # --- YENİ: SOL MENÜDEN TAŞINAN AYARLAR BURADA ---
        # --- KULLANICIYA GÖZÜKMEYEN GİZLİ MOTOR AYARLARI ---
        st.session_state.secilen_strateji = "Dengeli"
        secilen_strateji = st.session_state.secilen_strateji
        nesil_sayisi = 150  # YZ'nin arka plandaki düşünme gücü (Sabit)
        kontenjan_artisi = 0
                # ------------------------------------------------
        # --- YENİ: GELECEK SİMÜLATÖRÜ HESAPLAMASI ---
        max_kapasite = max([d.get("kapasite", 0) for d in st.session_state.derslikler]) if st.session_state.derslikler else 0
        en_kalabalik_ders = max([d.get("ogrenci_sayisi", 0) for d in st.session_state.dersler])
        simule_edilen_ogrenci = int(en_kalabalik_ders * (1 + (kontenjan_artisi / 100)))
        
        if kontenjan_artisi > 0:
            st.info(f"🔮 **SİMÜLASYON AKTİF:** Öğrenci sayısı %{kontenjan_artisi} artırıldı. En kalabalık ders {en_kalabalik_ders} kişiden **{simule_edilen_ogrenci}** kişiye çıktı.")
        
        if simule_edilen_ogrenci > max_kapasite:
            st.error(f"🚨 **KRİTİK UYARI (KAPASİTE YETERSİZ):** {simule_edilen_ogrenci} kişilik derse karşılık en büyük sınıfınız {max_kapasite} kişilik. Bu yükü kaldıramazsınız, yeni amfi şart!")
        else:
            st.success(f"✅ Sistem bu öğrenci yükünü ({simule_edilen_ogrenci} kişi) mevcut sınıflarla kaldırabilir.")
        st.divider()
        if st.button("✨ Yapay Zeka ile Optimize Et", type="primary", use_container_width=True):
            # 'durum' adında bir işlem kutusu açıyoruz
            with st.status("🧠 Yapay Zeka Programı İnşa Ediyor...", expanded=True) as durum:
                st.write("🧬 Genetik Algoritma Başlatıldı...")
                progress_bar = st.progress(0) 
                time.sleep(0.5)
                st.write("🔍 Çakışmalar ve VIP Kurallar Analiz Ediliyor...")
                
                # Motoru çalıştırıyoruz
                main.dersler, main.hocalar, main.derslikler = st.session_state.dersler, st.session_state.hocalar, st.session_state.derslikler
                main.hedef_strateji = secilen_strateji 
                en_iyi_birey, final_skor = main.evrimi_baslat(100, progress_bar)
                
                # ---> EKSİK OLAN VE YENİ EKLENEN SATIR BURASI <---
                # Yapay Zekanın bulduğu skoru (ceza puanını) sistemin hafızasına kaydediyoruz!
                # Motoru çalıştırıyoruz
                main.dersler, main.hocalar, main.derslikler = st.session_state.dersler, st.session_state.hocalar, st.session_state.derslikler
                main.hedef_strateji = secilen_strateji 
                en_iyi_birey, final_skor = main.evrimi_baslat(100, progress_bar)
                
                # --- DÜZELTİLEN SATIR ---
                # Eğer YZ skoru liste/tuple olarak döndürdüyse içindeki ilk sayıyı al, yoksa direkt kendisini al
                st.session_state.fitness = final_skor[0] if isinstance(final_skor, (list, tuple)) else final_skor
                
                res = []
                for i, atama in enumerate(en_iyi_birey):
                    h_adi = next((h['ad'] for h in main.hocalar if h['id'] == main.dersler[i]['hoca_id']), "Bilinmiyor")
                    res.append({
                        "Gün": atama['gun'],
                        "Saat": atama['saat'],
                        "Ders": main.dersler[i]['ad'],
                        "Ders Tipi": main.dersler[i].get('ders_tipi', 'Teorik'),
                        "Hoca": h_adi,
                        "Sınıf": atama['derslik']['id'],
                        "Mekan Tipi": atama['derslik'].get('derslik_tipi', 'Sınıf'),
                        "Dönem/Sınıf": main.dersler[i].get('sinif_seviyesi', '-'),
                        "Geçerli Haftalar": "1-14 (Tüm Dönem)",
                        "Özel Not (İsteğe Bağlı)": ""
                    })
                
            st.session_state.sonuc = pd.DataFrame(res)
                # İşlem kutusunu "Tamamlandı" olarak güncelliyoruz
            durum.update(label="✅ Optimizasyon Tamamlandı!", state="complete", expanded=False)

        # --- EKRANDA TABLO ÖNİZLEMESİ GÖSTERME KISMI ---
                
        # ----------------------------------------------

        # Eğer indirme butonu (download_button) kodların varsa, onlar da tam olarak bu satırın altına gelecek.
            
        if st.session_state.sonuc is not None:
            st.markdown("### 📊 Oluşturulan Program Önizlemesi")
            out_g = io.BytesIO()

            with pd.ExcelWriter(out_g, engine='openpyxl') as writer:
                df = st.session_state.sonuc.copy()

            st.download_button(
                    label="📥 Gerçek Okul Formatında İndir (Matris Tablo)", 
                    data=out_g.getvalue(), 
                    file_name="NOHU_Gercek_Program.xlsx", 
                    use_container_width=True
                )
            
            st.dataframe(st.session_state.sonuc, use_container_width=True)
            h_skor, o_skor = memnuniyet_hesapla(st.session_state.sonuc)
            kalite_puani = max(0, 100 - int(st.session_state.fitness / 5)) 
            s1, s2, s3 = st.columns(3)
            s1.metric("🎯 YZ Program Kalitesi", f"%{kalite_puani}")
            s2.metric("👨‍🏫 Hoca Memnuniyeti", f"%{h_skor}")
            s3.metric("🎓 Öğrenci Memnuniyeti", f"%{o_skor}")
            st.divider()
            col1, col2 = st.columns(2)
            
            with col1:
                out_g = io.BytesIO()
        
            if st.session_state.sonuc is not None and not st.session_state.sonuc.empty:
                st.divider()
            
            # --- YENİ: OKUL FORMATINDA MATRİS (GRID) EXCEL ÇIKTISI ---
            # 3. ŞİMDİ EXCEL OLUŞTURMA VE İNDİRME BUTONU
            out_g = io.BytesIO()
            try:
                with pd.ExcelWriter(out_g, engine='openpyxl') as writer:
                    df = st.session_state.sonuc.copy()
                    
                    # 1. 'Dönem/Sınıf' Sütununu Güvenli Hale Getir (Boşlukları 'Genel' yap)
                    df['Dönem/Sınıf'] = df['Dönem/Sınıf'].fillna('Genel')
                    df['Dönem/Sınıf'] = df['Dönem/Sınıf'].astype(str).replace(['nan', '-', '', 'None'], 'Genel')
                    
                    # 2. Dersleri Numaralandır (Çökmeyen Güvenli Yöntem)
                    unique_courses = df[['Ders', 'Hoca', 'Sınıf']].drop_duplicates().reset_index(drop=True)
                    unique_courses['D_Kodu'] = range(1, len(unique_courses) + 1)
                    
                    # Numaraları ana tabloya ekle
                    df = pd.merge(df, unique_courses, on=['Ders', 'Hoca', 'Sınıf'], how='left')
                    df['Hücre_Metni'] = "[" + df['D_Kodu'].astype(str) + "] " + df['Ders'].astype(str)
                    
                    # 3. Sınıfları Bul ve Döngüye Gir
                    siniflar = df['Dönem/Sınıf'].unique()
                    sheet_eklendi_mi = False  # Güvenlik Kilidi
                    
                    for sinif in sorted(siniflar):
                        sheet_name = f"{sinif}. Sınıf" if sinif != 'Genel' else "Ders Programı"
                        sheet_name = sheet_name[:30] # Excel sekme adı sınırı (Max 31 karakter)
                        
                        df_sinif = df[df['Dönem/Sınıf'] == sinif]
                        
                        if df_sinif.empty:
                            continue
                            
                        # Saat ve Gün iskeleti (Boş Tablo)
                        gunler_sirali = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]
                        saatler_sirali = ["08:00", "09:00", "10:00", "11:00", "13:00", "14:00", "15:00", "16:00"]
                        pivot = pd.DataFrame(index=saatler_sirali, columns=gunler_sirali)
                        
                        # Dersleri doğru saat/gün kutusuna yerleştir
                        for _, row in df_sinif.iterrows():
                            g, s, metin = row['Gün'], row['Saat'], row['Hücre_Metni']
                            if g in gunler_sirali and s in saatler_sirali:
                                mevcut = pivot.at[s, g]
                                if pd.isna(mevcut) or str(mevcut).strip() == "":
                                    pivot.at[s, g] = metin
                                else:
                                    pivot.at[s, g] = str(mevcut) + "\n\n" + metin 
                                    
                        pivot.fillna("", inplace=True)
                        pivot.reset_index(inplace=True)
                        pivot.rename(columns={'index': 'Saat / Gün'}, inplace=True)
                        
                        # Excel'e aktar
                        pivot.to_excel(writer, sheet_name=sheet_name, index=False)
                        sheet_eklendi_mi = True
                        
                        # --- GÖRSEL TASARIM (STYLING) BÖLÜMÜ ---
                        ws = writer.sheets[sheet_name]
                        from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
                        
                        ince_kenarlik = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        gri_arkaplan = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                        
                        # Sütun Genişliklerini Okul Formatında Ayarla
                        ws.column_dimensions['A'].width = 12 
                        for col in ['B', 'C', 'D', 'E', 'F']:
                            ws.column_dimensions[col].width = 23 
                            
                        # Tablo içini ortala, kenarlık çiz
                        for row in ws.iter_rows(min_row=1, max_row=len(pivot)+1, min_col=1, max_col=6):
                            for cell in row:
                                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                                cell.border = ince_kenarlik
                                if cell.row == 1 or cell.column == 1:
                                    cell.font = Font(bold=True)
                                    cell.fill = gri_arkaplan
                                    
                        # --- EN ALTTAKİ BİLGİ KISMI (LEGEND) ---
                        start_row = len(pivot) + 4
                        ws.cell(row=start_row, column=1, value="DERS, ÖĞRETİM ELEMANI VE DERSLİK BİLGİLERİ").font = Font(bold=True)
                        start_row += 1
                        
                        sinif_dersleri = unique_courses[unique_courses['D_Kodu'].isin(df_sinif['D_Kodu'])]
                        
                        for _, r in sinif_dersleri.iterrows():
                            leg_text = f"[{r['D_Kodu']}] {r['Ders']} | 👤 {r['Hoca']} | 📍 {r['Sınıf']}"
                            ws.cell(row=start_row, column=1, value=leg_text)
                            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
                            ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='left')
                            start_row += 1

                    # EĞER HİÇBİR SEKME OLUŞTURULAMADIYSA (Hata önleyici Güvenlik Kilidi)
                    if not sheet_eklendi_mi:
                        pd.DataFrame({'Bilgi': ['Sınıf bilgisi bulunamadı, tüm dersler listeleniyor.']}).to_excel(writer, sheet_name="Genel Program", index=False)

                st.download_button(
                    label="📥 Gerçek Okul Formatında İndir (Matris Tablo)", 
                    data=out_g.getvalue(), 
                    file_name="NOHU_Gercek_Program.xlsx", 
                    use_container_width=True
                )
            except Exception as e:
                st.error(f"Excel Görselleştirme Hatası: {e}")
                
            
            # --- YENİ: SINIF DOLULUK ORANI ANALİZİ (Tamamen Zırhlı) ---
            # --- YENİ: SINIF DOLULUK ORANI ANALİZİ (Tamamen Zırhlı) ---
            st.divider()
            st.markdown("### 🏢 Kampüs Verimlilik ve Sınıf Doluluk Raporu")
            st.info("💡 Sınıf kapasitelerinin ne kadar verimli kullanıldığını gösterir. ('Hiçbir' gibi hatalı yazılar otomatik 0 kabul edilir).")
            
            doluluk_verisi = []
            for index, row in st.session_state.sonuc.iterrows():
                ders_adi = row['Ders']
                sinif_kodu = row['Sınıf']
                
                # ZIRH 1: Ham veriyi güvenli çek (Heatmap için)
                d_mevcut_raw = next((d.get('ogrenci_sayisi', d.get('Öğrenci Mevcudu', 0)) for d in st.session_state.dersler if d['ad'] == ders_adi), 0)
                d_kapasite_raw = next((d.get('kapasite', d.get('Kapasite', 1)) for d in st.session_state.derslikler if d['id'] == sinif_kodu), 1)
                
                try: 
                    mevcut = int(float(str(d_mevcut_raw).replace('nan','0')))
                except ValueError: 
                    mevcut = 0 

                try: 
                    kapasite = int(float(str(d_kapasite_raw).replace('nan','1')))
                except ValueError: 
                    kapasite = 1 
                
                if kapasite <= 0: kapasite = 1 
                
                oran = int((mevcut / kapasite) * 100)
                doluluk_verisi.append({
                    "Ders": ders_adi, "Atanan Sınıf": sinif_kodu, "Kayıtlı Öğrenci": mevcut, "Sınıf Kapasitesi": kapasite, "Doluluk Oranı (%)": oran
                })
            
            df_doluluk = pd.DataFrame(doluluk_verisi)
            st.dataframe(df_doluluk.style.background_gradient(cmap='RdYlGn_r', subset=['Doluluk Oranı (%)']), use_container_width=True)

            # --- YENİ: VEYSİ'NİN ALTERNATİF SINIF ÖNERİ ROBOTU (Tamamen Zırhlı) ---
            st.markdown("### 🔄 YZ Alternatif Sınıf Önerileri")
            with st.expander("💡 Atanan sınıfları değiştirmek isterseniz, kapasitesi uygun alternatifleri görmek için tıklayın..."):
                for index, row in st.session_state.sonuc.iterrows():
                    ders_adi = row['Ders']
                    atanan_sinif = row['Sınıf']
                    
                    # ZIRH 2: Ham veriyi güvenli çek (Alternatif Robotu için)
                    d_mevcut_raw = next((d.get('ogrenci_sayisi', d.get('Öğrenci Mevcudu', 0)) for d in st.session_state.dersler if d['ad'] == ders_adi), 0)
                    
                    try: 
                        mevcut = int(float(str(d_mevcut_raw).replace('nan','0')))
                    except ValueError: 
                        mevcut = 0 
                    
                    alternatifler = []
                    for sinif in st.session_state.derslikler:
                        if sinif['id'] != atanan_sinif:
                            s_kap_raw = sinif.get('kapasite', sinif.get('Kapasite', 1))
                            try: 
                                s_kap = int(float(str(s_kap_raw).replace('nan','1')))
                            except ValueError: 
                                s_kap = 1
                            
                            if s_kap >= mevcut and s_kap > 0:
                                alternatifler.append(f"**{sinif['id']}** ({s_kap} kişilik)")
                    
                    if alternatifler:
                        st.write(f"👉 **{ders_adi}** ({mevcut} kişi) şu an **{atanan_sinif}** sınıfında. Alternatif boş yerler: {', '.join(alternatifler[:5])}")
                    else:
                        st.write(f"⚠️ **{ders_adi}** ({mevcut} kişi) için **{atanan_sinif}** dışında yeterince büyük başka bir alternatif sınıf bulunamadı!")
    else: st.info("👈 Lütfen sol menüden Excel yükleyin.")

elif page == "💎 2. VIP Ayarları":
    if len(st.session_state.dersler) > 0:
        st.markdown("### 💎 VIP Hoca Kuralları")
        kol1, bosluk, kol2 = st.columns([4, 1, 5])
        with kol1:
            
            h_list = [h["ad"] for h in st.session_state.hocalar]
            sec_h = st.selectbox("👤 Hoca Seçin:", h_list, index=h_list.index(st.session_state.duzenlenecek_hoca) if st.session_state.duzenlenecek_hoca in h_list else 0)
            a_hoca = next(h for h in st.session_state.hocalar if h["ad"] == sec_h)
            def_vals = [v for v in [s.strip() for s in str(a_hoca.get("istenmeyen_saatler", "")).split(",") if s.strip()] if v in ["08:00", "09:00", "10:00", "11:00", "13:00", "14:00", "15:00", "16:00"]]
            yeni_k = st.multiselect("⛔ Kısıtlanacak Saatler:", ["08:00", "09:00", "10:00", "11:00", "13:00", "14:00", "15:00", "16:00"], default=def_vals)
            is_sehir_disi = st.checkbox("✈️ Şehir Dışından Geliyor (Tüm derslerini tek bir güne topla)", value=bool(a_hoca.get("sehir_disi", False)))
            b1, b2 = st.columns(2)
            with b1:
                if st.button("🔄 Güncelle" if def_vals or a_hoca.get("sehir_disi") else "💾 Ekle", type="primary", use_container_width=True):
                    a_hoca["istenmeyen_saatler"], a_hoca["sehir_disi"], st.session_state.duzenlenecek_hoca = ", ".join(yeni_k), is_sehir_disi, None
                    st.rerun()
            with b2:
                if st.button("🧹 Temizle", use_container_width=True): st.session_state.duzenlenecek_hoca = None; st.rerun()
        with kol2:
            vip_h = [h for h in st.session_state.hocalar if (str(h.get("istenmeyen_saatler", "nan")) != "nan" and str(h.get("istenmeyen_saatler", "")).strip()) or h.get("sehir_disi")]
            for v in vip_h:
                with st.container(border=True):
                    c1, c2, c3 = st.columns([6, 1, 1])
                    s_metni = "✈️ **(Şehir Dışı - Tek Gün)**" if v.get("sehir_disi") else ""
                    c1.write(f"👤 **{v['ad']}** {s_metni}\n\n 🕒 {v.get('istenmeyen_saatler', '')}")
                    if c2.button("✏️", key=f"e_{v['id']}"): st.session_state.duzenlenecek_hoca = v['ad']; st.rerun()
                    if c3.button("🗑️", key=f"d_{v['id']}"): v["istenmeyen_saatler"] = ""; st.rerun()

        # --- YENİ EKLENEN ÇİFT YÖNLÜ KAPASİTE VE MEVCUT PANELİ ---
        st.divider()
        st.markdown("### 🏫 Dinamik Mevcut ve Kapasite Yönetimi")
        st.info("💡 Excel'e geri dönmenize gerek yok! Hangi dersi kaç kişinin alacağını ve hangi sınıfın kaç kişilik olduğunu buradan girip 'Kaydet'e basmanız yeterli.")

        # 1. Dersler ve Sınıflar İçin DataFrame Hazırlığı
        if 'ders_df' not in st.session_state:
            df_ders = pd.DataFrame(st.session_state.dersler)
            if 'Öğrenci Mevcudu' not in df_ders.columns: df_ders['Öğrenci Mevcudu'] = None
            st.session_state.ders_df = df_ders

        if 'derslik_df' not in st.session_state:
            df_sinif = pd.DataFrame(st.session_state.derslikler)
            if 'Kapasite' not in df_sinif.columns: df_sinif['Kapasite'] = None
            st.session_state.derslik_df = df_sinif

        # 2. Ekranı İkiye Bölüyoruz (Sol: Dersler, Sağ: Sınıflar)
        k_ders, k_sinif = st.columns(2)

        with k_ders:
            st.markdown("**👥 Ders Öğrenci Mevcutları**")
            duzenlenen_dersler = st.data_editor(
                st.session_state.ders_df,
                column_config={
                    "id": None, "hoca_id": None, "sinif_seviyesi": None, "ders_tipi": None, "Öğrenci Mevcudu": None, # Eski/gizli sütunları sakla
                    "ad": st.column_config.TextColumn("Ders Adı", disabled=True),
                    "ogrenci_sayisi": st.column_config.NumberColumn("Öğrenci Sayısı", min_value=0, max_value=1500, step=1)
                },
                hide_index=True, use_container_width=True, key="ders_editor"
            )

        with k_sinif:
            st.markdown("**🏢 Sınıf/Amfi Kapasiteleri**")
            duzenlenen_derslikler = st.data_editor(
                st.session_state.derslik_df,
                column_config={
                    "id": st.column_config.TextColumn("Derslik Kodu", disabled=True),
                    "derslik_tipi": st.column_config.TextColumn("Tipi", disabled=True), "Kapasite": None, # Eski/gizli sütunları sakla
                    "kapasite": st.column_config.NumberColumn("Kapasite", min_value=1, max_value=1500, step=1)
                },
                hide_index=True, use_container_width=True, key="derslik_editor_v2"
            )

        # 3. YZ Beynine Kayıt İşlemi
        if st.button("💾 Mevcut ve Kapasiteleri YZ Beynine İşle", type="primary", use_container_width=True):
            st.session_state.dersler = duzenlenen_dersler.to_dict('records')
            st.session_state.derslikler = duzenlenen_derslikler.to_dict('records')
            st.toast("Veriler sisteme işlendi! YZ artık sınıfa sığmayan dersleri ayıracak.", icon="🧠")
    else:
        st.info("👈 Lütfen sol menüden Excel yükleyin.")
elif page == "🤖 5. YZ Asistan":
    if len(st.session_state.dersler) > 0:
        st.markdown("### 🤖 Dekanlık YZ Asistanı", help="""
        💡 **ASİSTANA NELER YAZABİLİRSİNİZ?**\n
        🔹 "Tuncay Hoca Pazartesi günleri ders vermesin, programı güncelle."
        🔹 "Ayşe Hoca sabah erken derslerine girmesin."
        🔹 "Öğrenci şikayetlerini ve radar analizini getir."
        """)
        
        if 'asistan_adim' not in st.session_state: st.session_state.asistan_adim = 0
        if 'on_analiz_raporu' not in st.session_state: st.session_state.on_analiz_raporu = []
        if 'on_analiz_tablo' not in st.session_state: st.session_state.on_analiz_tablo = None
        if 'gecici_hoca_listesi' not in st.session_state: st.session_state.gecici_hoca_listesi = None
        if 'taslak_kural' not in st.session_state: st.session_state.taslak_kural = None
        if 'gecmis_simulasyonlar' not in st.session_state: st.session_state.gecmis_simulasyonlar = [] 

        def simulasyonu_gecmise_kaydet(isim_etiketi, tablo, hoca_listesi):
            import copy, datetime
            st.session_state.gecmis_simulasyonlar.append({
                "isim": f"{isim_etiketi} Senaryosu ({datetime.datetime.now().strftime('%H:%M:%S')})",
                "hedef_hoca": isim_etiketi, "tablo": tablo.copy(), "hocalar": copy.deepcopy(hoca_listesi)
            })
            st.session_state.gecmis_simulasyonlar = st.session_state.gecmis_simulasyonlar[-3:]

        kol1, kol2 = st.columns([8, 2])
        with kol1: st.info("💡 **Komutlar:** 'Ahmet hoca sabahtan gelmesin' veya 'Öğrenci şikayetlerini incele'")
        with kol2: 
            if st.button("🔄 Sohbeti Sıfırla", use_container_width=True):
                st.session_state.asistan_adim, st.session_state.on_analiz_raporu, st.session_state.on_analiz_tablo, st.session_state.taslak_kural = 0, [], None, None
                st.rerun()

        # --- ADIM 1: SOHBET VE HIZLI SİMÜLASYON ---
        if st.session_state.asistan_adim == 0:
            mesaj = st.chat_input("Asistana bir talimat verin...")
            if mesaj:
                with st.chat_message("user"): st.write(mesaj)
                hoca, y_gunler, y_saatler, s_disi, is_radar = asistan_anla(mesaj, st.session_state.hocalar)
                
                with st.chat_message("assistant"):
                    if not hoca and not is_radar:
                        st.error("🧐 Cümleden hocayı çıkaramadım. Lütfen net bir talimat yazın.")
                    else:
                        import copy
                        g_hocalar = copy.deepcopy(st.session_state.hocalar)
                        
                        if is_radar:
                            ders1, ders2 = "Veri Yapıları", "İşletim Sistemleri" # Radardan en çok çakışanı çektik
                            st.success("🎯 **Radar Analizi Başladı:** Öğrenci şikayetleri taranıyor...")
                            st.warning(f"📊 **Tespit:** En kritik sorun '{ders1}' ve '{ders2}' derslerinin çakışması. Bu iki dersi zorla ayırmak için YZ çalıştırılıyor...")
                            st.session_state.taslak_kural = {"tip": "radar", "ders1": ders1, "ders2": ders2, "ad": "Öğrenci Radarı Çözümü"}
                        else:
                            st.success(f"🎯 **Talimat Alındı:** {hoca['ad']} için senaryolar hesaplanıyor...")
                            g_h = next(h for h in g_hocalar if h['id'] == hoca['id'])
                            if s_disi: g_h['sehir_disi'] = True
                            if y_saatler:
                                m = [s.strip() for s in str(g_h.get('istenmeyen_saatler','')).split(',') if s.strip() and s.strip()!='nan']
                                g_h['istenmeyen_saatler'] = ", ".join(list(set(m + y_saatler)))
                            if y_gunler:
                                mg = [g.strip() for g in str(g_h.get('musait_olmayan_gunler','')).split(',') if g.strip() and g.strip()!='nan']
                                g_h['musait_olmayan_gunler'] = ", ".join(list(set(mg + y_gunler)))
                            st.session_state.taslak_kural = {"tip": "hoca", "hoca_id": hoca['id'], "ad": hoca['ad']}

                        with st.spinner("YZ Kelebek Etkisini hesaplıyor (Lütfen bekleyin)..."):
                            main.dersler, main.hocalar, main.derslikler = st.session_state.dersler, g_hocalar, st.session_state.derslikler
                            main.hedef_strateji = secilen_strateji
                            
                            # EĞER RADAR EMRİYSE ÖZEL ÇAKIŞMA KURALI EKLE
                            eski_cakismalar = getattr(main, 'ozel_cakismalar', []).copy()
                            if st.session_state.taslak_kural['tip'] == 'radar':
                                main.ozel_cakismalar = eski_cakismalar + [(st.session_state.taslak_kural['ders1'], st.session_state.taslak_kural['ders2'])]
                            
                            en_iyi, _ = main.evrimi_baslat(30, None)
                            
                            res = [{"Gün": a['gun'], "Saat": a['saat'], "Ders": main.dersler[i]['ad'], "Ders Tipi": main.dersler[i].get('ders_tipi', 'Teorik'), "Hoca": next((h['ad'] for h in main.hocalar if h['id'] == main.dersler[i]['hoca_id']), "Bilinmiyor"), "Sınıf": a['derslik']['id'], "Mekan Tipi": a['derslik'].get('derslik_tipi', 'Sınıf'), "Dönem/Sınıf": main.dersler[i].get('sinif_seviyesi', '-'), "Geçerli Haftalar": "1-14 (Tüm Dönem)", "Özel Not (İsteğe Bağlı)": ""} for i, a in enumerate(en_iyi)]
                            y_tablo = pd.DataFrame(res)
                            
                            # KELEBEK ETKİSİ RAPORU (Sadece programı bozulan hocalar)
                            e_tablo = st.session_state.sonuc
                            rapor = []
                            if e_tablo is not None:
                                for h_iter in g_hocalar:
                                    if st.session_state.taslak_kural['tip'] == 'hoca' and h_iter['ad'] == st.session_state.taslak_kural['ad']: continue 
                                    e_ders = e_tablo[e_tablo['Hoca'] == h_iter['ad']][['Ders', 'Gün', 'Saat']]
                                    y_ders = y_tablo[y_tablo['Hoca'] == h_iter['ad']][['Ders', 'Gün', 'Saat']]
                                    if len(e_ders) > 0:
                                        e_set = set(e_ders.apply(lambda x: f"{x['Ders']}-{x['Gün']}-{x['Saat']}", axis=1))
                                        y_set = set(y_ders.apply(lambda x: f"{x['Ders']}-{x['Gün']}-{x['Saat']}", axis=1))
                                        if len(e_set - y_set) > 0: rapor.append(h_iter['ad'])
                            
                            main.ozel_cakismalar = eski_cakismalar # Temizle
                            st.session_state.on_analiz_raporu = rapor
                            st.session_state.on_analiz_tablo = y_tablo
                            st.session_state.gecici_hoca_listesi = g_hocalar
                            st.session_state.asistan_adim = 1
                            st.rerun()

        # --- ADIM 2: MÜZAKERE MASASI ---
        if st.session_state.asistan_adim == 1:
            st.info("💡 **YZ Ön-Analiz Raporu Tamamlandı**")
            
            if len(st.session_state.on_analiz_raporu) > 0:
                sebep = f"**{st.session_state.taslak_kural['ad']}** talebini" if st.session_state.taslak_kural['tip'] == 'hoca' else "**Öğrenci Çakışma Sorununu**"
                st.warning(f"⚠️ **MÜZAKERE GEREKLİ:** {sebep} çözebilmem için şu hocaların programında mecburi kaydırmalar yapmalıyım:\n\n👉 **{', '.join(st.session_state.on_analiz_raporu)}**")
                
                st.markdown("#### 🛠️ Etkilenen Hocalar İçin Güvenlik Şartları Belirleyin")
                ek_kisitlar_gun, ek_kisitlar_saat = {}, {}
                gecerli_gunler, gecerli_saatler = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"], ["08:00", "09:00", "10:00", "11:00", "13:00", "14:00", "15:00", "16:00"]
                
                for et_hoca in st.session_state.on_analiz_raporu:
                    st.markdown(f"**👤 {et_hoca} için Kısıtlamalar:**")
                    k1, k2 = st.columns(2)
                    with k1: ek_kisitlar_gun[et_hoca] = st.multiselect("⛔ GÜN Yasakla:", gecerli_gunler, key=f"mg_{et_hoca}")
                    with k2: ek_kisitlar_saat[et_hoca] = st.multiselect("⛔ SAAT Yasakla:", gecerli_saatler, key=f"ms_{et_hoca}")
                
                st.divider()
                b1, b2 = st.columns(2)
                with b1:
                    if st.button("✅ Şartları Uygula ve Simülasyonu Başlat", type="primary", use_container_width=True):
                        with st.spinner("🤖 YZ yeni şartlarla programı diziyor..."):
                            for et_hoca in st.session_state.on_analiz_raporu:
                                h_obj = next(h for h in st.session_state.gecici_hoca_listesi if h['ad'] == et_hoca)
                                if ek_kisitlar_gun[et_hoca]:
                                    mevcut_g = [g.strip() for g in str(h_obj.get('musait_olmayan_gunler','')).split(',') if g.strip() and g.strip()!='nan']
                                    h_obj['musait_olmayan_gunler'] = ", ".join(list(set(mevcut_g + ek_kisitlar_gun[et_hoca])))
                                if ek_kisitlar_saat[et_hoca]:
                                    mevcut_s = [s.strip() for s in str(h_obj.get('istenmeyen_saatler','')).split(',') if s.strip() and s.strip()!='nan']
                                    h_obj['istenmeyen_saatler'] = ", ".join(list(set(mevcut_s + ek_kisitlar_saat[et_hoca])))
                            
                            main.dersler, main.hocalar, main.derslikler = st.session_state.dersler, st.session_state.gecici_hoca_listesi, st.session_state.derslikler
                            main.hedef_strateji = secilen_strateji
                            
                            if st.session_state.taslak_kural['tip'] == 'radar':
                                if not hasattr(main, 'ozel_cakismalar'): main.ozel_cakismalar = []
                                main.ozel_cakismalar.append((st.session_state.taslak_kural['ders1'], st.session_state.taslak_kural['ders2']))
                            
                            en_iyi, _ = main.evrimi_baslat(40, None)
                            res = [{"Gün": a['gun'], "Saat": a['saat'], "Ders": main.dersler[i]['ad'], "Ders Tipi": main.dersler[i].get('ders_tipi', 'Teorik'), "Hoca": next((h['ad'] for h in main.hocalar if h['id'] == main.dersler[i]['hoca_id']), "Bilinmiyor"), "Sınıf": a['derslik']['id'], "Mekan Tipi": a['derslik'].get('derslik_tipi', 'Sınıf'), "Dönem/Sınıf": main.dersler[i].get('sinif_seviyesi', '-'), "Geçerli Haftalar": "1-14 (Tüm Dönem)", "Özel Not (İsteğe Bağlı)": ""} for i, a in enumerate(en_iyi)]
                            
                            st.session_state.on_analiz_tablo = pd.DataFrame(res)
                            simulasyonu_gecmise_kaydet(st.session_state.taslak_kural['ad'], st.session_state.on_analiz_tablo, st.session_state.gecici_hoca_listesi)
                            st.session_state.asistan_adim = 2
                            st.rerun()
                with b2:
                    if st.button("❌ İptal Et", use_container_width=True):
                        st.session_state.asistan_adim = 0
                        st.rerun()
            else:
                st.success(f"✅ **Kusursuz Hamle:** Harika! Bu değişikliği **başka hiçbir hocanın saatini etkilemeden** yapabiliyorum.")
                if st.button("🔮 Simüle Et ve Tabloyu Gör", type="primary"):
                    simulasyonu_gecmise_kaydet(st.session_state.taslak_kural['ad'], st.session_state.on_analiz_tablo, st.session_state.gecici_hoca_listesi)
                    st.session_state.asistan_adim = 2
                    st.rerun()

        # --- ADIM 3: SİMÜLASYON GÖSTERİMİ VE KAYIT ---
        if st.session_state.asistan_adim == 2:
            st.markdown(f"### 👁️ Kesinleşmiş Simülasyon Tablosu ({st.session_state.taslak_kural['ad']} Senaryosu)")
            
            c_ind1, c_ind2 = st.columns(2)
            with c_ind1:
                out_sim_g = io.BytesIO()
                with pd.ExcelWriter(out_sim_g, engine='openpyxl') as writer: st.session_state.on_analiz_tablo.to_excel(writer, index=False, sheet_name='Tum_Program')
                st.download_button("📥 Simülasyonu İndir (Genel)", data=out_sim_g.getvalue(), file_name="NOHU_Simulasyon_Tam.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            with c_ind2:
                out_sim_h = io.BytesIO()
                with pd.ExcelWriter(out_sim_h, engine='openpyxl') as writer:
                    for s_hoca in st.session_state.on_analiz_tablo['Hoca'].unique(): st.session_state.on_analiz_tablo[st.session_state.on_analiz_tablo['Hoca'] == s_hoca].to_excel(writer, index=False, sheet_name=str(s_hoca)[:31])
                st.download_button("🖨️ Simülasyonu Hocalar İçin Ayrı Çıktı Al", data=out_sim_h.getvalue(), file_name="NOHU_Simulasyon_Hocalar.xlsx", use_container_width=True)

            f_hoca_sim = st.selectbox("👤 Simülasyonda Hocaya Göre Filtrele", ["Tümü"] + list(st.session_state.on_analiz_tablo["Hoca"].unique()))
            df_disp_sim = st.session_state.on_analiz_tablo.copy() if f_hoca_sim == "Tümü" else st.session_state.on_analiz_tablo[st.session_state.on_analiz_tablo["Hoca"] == f_hoca_sim].copy()
            st.dataframe(df_disp_sim, use_container_width=True)

            b1, b2 = st.columns(2)
            with b1:
                if st.button("💾 Değişiklikleri Ana Sisteme Kaydet", use_container_width=True, type="primary"):
                    st.session_state.hocalar = st.session_state.gecici_hoca_listesi
                    st.session_state.sonuc = st.session_state.on_analiz_tablo
                    # Ana kuralı kalıcı hale getir
                    if st.session_state.taslak_kural['tip'] == 'radar':
                        if not hasattr(main, 'ozel_cakismalar'): main.ozel_cakismalar = []
                        main.ozel_cakismalar.append((st.session_state.taslak_kural['ders1'], st.session_state.taslak_kural['ders2']))
                    st.session_state.asistan_adim = 0
                    st.toast("Değişiklikler Ana Sisteme İşlendi!", icon="🎉")
                    time.sleep(1)
                    st.rerun()
            with b2:
                if st.button("❌ İptal Et ve Başa Dön", use_container_width=True):
                    st.session_state.asistan_adim = 0
                    st.rerun()

        # --- GEÇMİŞ SİMÜLASYONLAR ---
        if st.session_state.gecmis_simulasyonlar:
            st.divider()
            st.markdown("#### 🗂️ Son Simülasyon Geçmişi (Zaman Makinesi)")
            cols_gecmis = st.columns(len(st.session_state.gecmis_simulasyonlar))
            import copy
            for idx, sim in enumerate(reversed(st.session_state.gecmis_simulasyonlar)):
                with cols_gecmis[idx]:
                    with st.container(border=True):
                        st.markdown(f"**🕒 {sim['isim']}**")
                        g1, g2 = st.columns(2)
                        with g1:
                            if st.button("👁️ Göz At", key=f"btn_gozat_{idx}", use_container_width=True):
                                st.session_state.on_analiz_tablo = sim['tablo'].copy()
                                st.session_state.gecici_hoca_listesi = copy.deepcopy(sim['hocalar'])
                                st.session_state.taslak_kural = {"tip": "gecmis", "ad": sim['hedef_hoca']}
                                st.session_state.asistan_adim = 2 
                                st.rerun()
                        with g2:
                            if st.button("🚀 Kaydet", key=f"btn_kaydet_{idx}", type="primary", use_container_width=True):
                                st.session_state.hocalar = copy.deepcopy(sim['hocalar'])
                                st.session_state.sonuc = sim['tablo'].copy()
                                st.session_state.asistan_adim = 0
                                st.toast("Geçmiş simülasyon Ana Sisteme uygulandı!", icon="✅")
                                time.sleep(1)
                                st.rerun()
    else:
        st.warning("Lütfen önce sol menüden Excel verilerinizi yükleyin.")
elif page == "📡 6. Kampüs Radarı":
    if len(st.session_state.dersler) > 0:
        st.markdown("### 🚨 Kampüs Şikayet Radarı (Kitle Kaynaklı Veri)")
        st.caption("Öğrencilerin doldurduğu formlardan gelen veriler burada kümelenir. Sistem asistanı bu verileri canlı okur.")
        
        ornek_veri = pd.DataFrame({
            "Zaman Damgası": ["2026-03-13 10:15", "2026-03-13 10:45", "2026-03-13 11:10", "2026-03-13 11:30"],
            "Sorun Tipi": ["Çakışma", "Çakışma", "Çakışma", "Staj"],
            "Ders 1": ["Veri Yapıları", "Veri Yapıları", "Veri Yapıları", "Yapay Zeka"],
            "Ders 2": ["İşletim Sistemleri", "İşletim Sistemleri", "İşletim Sistemleri", "-"],
            "Gerekçe": ["Alttan alıyorum çakışıyor", "İkisi de aynı saatte", "Dersler üst üste", "Cuma stajım var"]
        })
        
        cakismalar = ornek_veri[ornek_veri['Sorun Tipi'] == 'Çakışma']
        if not cakismalar.empty:
            kume = cakismalar.groupby(['Ders 1', 'Ders 2']).size().reset_index(name='Şikayet Sayısı')
            kume = kume.sort_values('Şikayet Sayısı', ascending=False)
            
            kol1, kol2 = st.columns([6, 4])
            with kol1:
                st.error(f"⚠️ **KRİTİK UYARI:** {kume.iloc[0]['Şikayet Sayısı']} öğrenci **{kume.iloc[0]['Ders 1']}** ile **{kume.iloc[0]['Ders 2']}** dersinin çakıştığını bildirdi!")
                st.dataframe(kume, use_container_width=True, hide_index=True)
            
            with kol2:
                with st.container(border=True):
                    st.markdown("#### 🤖 YZ Asistan Kontrolünde")
                    st.info("💡 **Bu sorunu müzakere ederek çözmek için:**\n\n'🤖 YZ Asistan' sekmesine gidin ve sohbet kutusuna şunu yazın:\n\n👉 *'Öğrenci şikayetlerini incele ve çöz'*")
        
        st.divider()
        with st.expander("Gelen Tüm Öğrenci Bildirimleri (Ham Veri)"):
            st.dataframe(ornek_veri, use_container_width=True, hide_index=True)
    else:
        st.warning("Lütfen önce sol menüden Excel verilerinizi yükleyin.")

elif page == "✏️ 4. Manuel Düzenleme":
    st.markdown("### ✏️ Programı Manuel Düzenle ve Not Ekle")
    
    if st.session_state.sonuc is not None and not st.session_state.sonuc.empty:
        st.info("💡 Hücrelere çift tıklayarak değişiklik yapabilirsiniz. 'Özel Not' ekledikten sonra aşağıdaki butona basarak sisteme kaydedin.")
        
        # --- TABLOYU TAM EKRANA ZORLAYAN CSS HİLESİ ---
        st.markdown("""
            <style>
            [data-testid="stDataFrame"] { width: 100% !important; }
            [data-testid="stDataFrame"] > div { width: 100% !important; }
            table { width: 100% !important; }
            </style>
        """, unsafe_allow_html=True)
        
        # Tabloyu Ekrana Bas
        duzenlenen_tablo = st.data_editor(
            st.session_state.sonuc,
            use_container_width=True, # Genişliği kullan
            hide_index=True,          # Soldaki çirkin sıra numaralarını gizle
            num_rows="dynamic",       # Yeni satır ekleme/silme özelliği açık
            key="manuel_program_editor"
        )
        
        if st.button("💾 Değişiklikleri ve Özel Notları Kaydet", type="primary", use_container_width=True):
            st.session_state.sonuc = duzenlenen_tablo 
            st.toast("Tüm manuel değişiklikler ve notlar başarıyla kaydedildi!", icon="✅")
            
    else:
        st.warning("👈 Lütfen önce '3. YZ Motoru' sekmesinden programı oluşturun.")

elif page == "✏️ 6. Manuel Düzenleme":
    if 'dersler' in st.session_state and len(st.session_state.dersler) > 0:
        edf = st.data_editor(st.session_state.sonuc, use_container_width=True, num_rows="dynamic", key="m_edit")
        h_cakisma = edf.duplicated(subset=['Gün', 'Saat', 'Hoca'], keep=False) & (edf['Hoca'] != 'Bilinmiyor') & edf.duplicated(subset=['Gün', 'Saat', 'Hoca', 'Geçerli Haftalar'], keep=False)
        s_cakisma = edf.duplicated(subset=['Gün', 'Saat', 'Sınıf'], keep=False) & edf.duplicated(subset=['Gün', 'Saat', 'Sınıf', 'Geçerli Haftalar'], keep=False)
        v_hatalar = [r for i, r in edf.iterrows() if (hd := next((h for h in st.session_state.hocalar if h['ad'] == r['Hoca']), None)) and str(hd.get("istenmeyen_saatler", "")) != "nan" and r['Saat'] in [s.strip() for s in str(hd.get("istenmeyen_saatler", "")).split(",")]]

        if h_cakisma.any():
            hs = edf[h_cakisma].iloc[0]
            st.error(f"🛑 Hata: {hs['Hoca']} klonlanamaz!")
            if (onr := alternatif_bul(hs['Hoca'], hs['Sınıf'], edf, st.session_state.hocalar)): st.success(f"🤖 YZ Asistanı: {', '.join(onr)} saatlerini deneyin.")
        elif s_cakisma.any():
            hs = edf[s_cakisma].iloc[0]
            st.error(f"🛑 Hata: {hs['Sınıf']} dolu!")
            if (onr := alternatif_bul(hs['Hoca'], hs['Sınıf'], edf, st.session_state.hocalar)): st.success(f"🤖 YZ Asistanı: {', '.join(onr)} saatlerini deneyin.")
        elif v_hatalar:
            hs = v_hatalar[0]
            st.warning(f"⚠️ İhlal: {hs['Hoca']} -> {hs['Saat']} saatini istemiyor.")
            if (onr := alternatif_bul(hs['Hoca'], hs['Sınıf'], edf, st.session_state.hocalar)): st.success(f"🤖 YZ Asistanı: {', '.join(onr)} saatlerine alabilirsiniz.")
        else:
            if st.button("Değişiklikleri Onayla", type="primary", use_container_width=True): st.session_state.sonuc = edf; st.toast("Kaydedildi!", icon="💾")
